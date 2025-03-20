from flask import Blueprint, render_template, request, jsonify, send_file
from datetime import datetime, timedelta, date
from sqlalchemy import func, text, extract, and_, cast, Date
from models import db, Sale, Product, Category, Receivable, Payable, SaleItem
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from decimal import Decimal
import traceback
from dateutil.relativedelta import relativedelta

management = Blueprint('management', __name__)

@management.route('/painel')
@login_required
def dashboard():
    # Cálculo das vendas do dia
    today = datetime.now().date()
    daily_sales = db.session.query(func.sum(Sale.total)).filter(
        func.date(Sale.date) == today
    ).scalar() or 0

    # Cálculo das vendas do mês
    first_day = today.replace(day=1)
    monthly_sales = db.session.query(func.sum(Sale.total)).filter(
        func.date(Sale.date) >= first_day
    ).scalar() or 0

    # Contas a receber
    receivables = db.session.query(func.sum(Receivable.amount)).filter(
        Receivable.status == 'pending'
    ).scalar() or 0

    # Contas a pagar
    payables = db.session.query(func.sum(Payable.amount)).filter(
        Payable.status == 'pending'
    ).scalar() or 0

    # Dados para o gráfico de vendas dos últimos 7 dias
    dates = []
    sales_data = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        dates.append(date.strftime('%d/%m'))
        daily_total = db.session.query(func.sum(Sale.total)).filter(
            func.date(Sale.date) == date
        ).scalar() or 0
        sales_data.append(float(daily_total))

    # Top produtos vendidos
    top_products = db.session.query(
        Product.name,
        func.sum(SaleItem.quantity).label('quantity'),
        func.sum(SaleItem.subtotal).label('total')
    ).join(SaleItem).group_by(Product.id, Product.name).order_by(
        func.sum(SaleItem.subtotal).desc()
    ).limit(5).all()

    top_products = [
        {
            'name': p.name,
            'quantity': float(p.quantity),
            'total': float(p.total)
        } for p in top_products
    ]

    return render_template('management/dashboard.html',
                         daily_sales="{:.2f}".format(daily_sales),
                         monthly_sales="{:.2f}".format(monthly_sales),
                         receivables="{:.2f}".format(receivables),
                         payables="{:.2f}".format(payables),
                         dates=dates,
                         sales_data=sales_data,
                         top_products=top_products)

@management.route('/gestao')
@login_required
def gestao():
    try:
        # Dados para o gráfico de vendas dos últimos 12 meses
        today = datetime.now()
        last_12_months = []
        sales_by_month = []
        
        for i in range(11, -1, -1):
            date = today - relativedelta(months=i)
            start_date = date.replace(day=1, hour=0, minute=0, second=0)
            if date.month == 12:
                end_date = date.replace(year=date.year + 1, month=1, day=1, hour=0, minute=0, second=0)
            else:
                end_date = date.replace(month=date.month + 1, day=1, hour=0, minute=0, second=0)
            end_date = end_date - timedelta(seconds=1)
            
            sales = Sale.query.filter(
                Sale.date >= start_date,
                Sale.date <= end_date
            ).all()
            
            monthly_total_sales = sum(sale.total for sale in sales)
            last_12_months.append(date.strftime('%b/%Y'))
            sales_by_month.append(float(monthly_total_sales))
        
        # Top 5 produtos mais vendidos
        top_products = db.session.query(
            Product.name,
            func.sum(SaleItem.quantity).label('total_quantity'),
            func.sum(SaleItem.subtotal).label('total_value')
        ).join(SaleItem).group_by(Product.id, Product.name).order_by(text('total_quantity DESC')).limit(5).all()
        
        # Dados financeiros do mês atual
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0)
        if today.month == 12:
            end_of_month = today.replace(year=today.year + 1, month=1, day=1, hour=0, minute=0, second=0)
        else:
            end_of_month = today.replace(month=today.month + 1, day=1, hour=0, minute=0, second=0)
        end_of_month = end_of_month - timedelta(seconds=1)
        
        print(f"\nBuscando dados do mês atual...")
        print(f"Data inicial: {start_of_month}")
        print(f"Data final: {end_of_month}")
        
        # Vendas do mês
        month_sales = Sale.query.filter(
            Sale.date >= start_of_month,
            Sale.date <= end_of_month
        ).all()
        month_revenue = sum(sale.total for sale in month_sales)
        print(f"Total de vendas do mês: {len(month_sales)}")
        print(f"Receita do mês: {month_revenue}")
        
        # Vendas do dia (total_sales)
        today_date = datetime.now().date()
        day_sales = Sale.query.filter(
            func.date(Sale.date) == today_date
        ).all()
        total_sales = sum(sale.total for sale in day_sales)
        print(f"Total de vendas do dia: {total_sales}")
        
        # Contas a receber do mês
        print("\nBuscando contas a receber...")
        try:
            # Primeiro vamos ver todas as contas a receber
            all_receivables = Receivable.query.all()
            print(f"Total de contas a receber no banco: {len(all_receivables)}")
            
            # Filtrando pelo mês atual
            month_receivables = Receivable.query.filter(
                Receivable.due_date >= start_of_month,
                Receivable.due_date <= end_of_month
            ).all()
            
            print("\nContas a receber do mês:")
            for r in month_receivables:
                print(f"ID: {r.id}")
                print(f"Vencimento: {r.due_date}")
                print(f"Valor: {float(r.amount)}")
                print(f"Valor Pago: {float(r.paid_amount) if r.paid_amount else 0}")
                print(f"Valor Restante: {float(r.remaining_amount) if r.remaining_amount else float(r.amount)}")
                print("---")
            
            # Calculando os totais - atualizando para usar todas as contas a receber
            total_receivables = sum(float(r.amount) for r in all_receivables if r.status in ['pending', 'partial', 'overdue'])
            received = sum(float(r.paid_amount if r.paid_amount else 0) for r in all_receivables if r.status in ['pending', 'partial', 'overdue', 'paid'])
            pending_receivables = sum(float(r.remaining_amount if r.remaining_amount else r.amount) for r in all_receivables if r.status in ['pending', 'partial', 'overdue'])
            
            print(f"\nTotais calculados:")
            print(f"Total a receber: {total_receivables}")
            print(f"Recebido: {received}")
            print(f"Pendente: {pending_receivables}")
            
        except Exception as e:
            print(f"Erro ao processar contas a receber: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            total_receivables = 0
            received = 0
            pending_receivables = 0
        
        # Contas a pagar do mês
        print("\nBuscando contas a pagar...")
        try:
            # Primeiro vamos ver todas as contas a pagar
            all_payables = Payable.query.all()
            print(f"Total de contas a pagar no banco: {len(all_payables)}")
            
            # Filtrando pelo mês atual
            month_payables = Payable.query.filter(
                Payable.due_date >= start_of_month,
                Payable.due_date <= end_of_month
            ).all()
            print(f"Total de contas a pagar do mês: {len(month_payables)}")
            
            # Calculando os totais - atualizando para usar os valores corretos
            total_payables = sum(float(p.amount) for p in all_payables if p.status in ['pending', 'partial', 'overdue'])
            paid = sum(float(p.paid_amount) for p in all_payables if p.status in ['pending', 'partial', 'overdue', 'paid'])
            pending_payables = sum(float(p.remaining_amount) for p in all_payables if p.status in ['pending', 'partial', 'overdue'])
            
            print(f"Total a pagar: {total_payables}")
            print(f"Pago: {paid}")
            print(f"Pendente: {pending_payables}")
            
        except Exception as e:
            print(f"Erro ao processar contas a pagar: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            total_payables = 0
            paid = 0
            pending_payables = 0
        
        # Cálculo do lucro estimado
        estimated_profit = float(month_revenue) - float(total_payables)
        
        # Status dos recebíveis
        overdue_receivables = Receivable.query.filter(
            Receivable.due_date < today,
            Receivable.remaining_amount > 0
        ).all()
        total_overdue = sum(float(r.remaining_amount) for r in overdue_receivables)
        
        # Tendência de crescimento (comparação com mês anterior)
        last_month_start = start_of_month - relativedelta(months=1)
        last_month_end = start_of_month - timedelta(seconds=1)
        
        last_month_sales = Sale.query.filter(
            Sale.date >= last_month_start,
            Sale.date <= last_month_end
        ).all()
        last_month_revenue = sum(sale.total for sale in last_month_sales)
        
        if last_month_revenue > 0:
            growth = ((month_revenue - last_month_revenue) / last_month_revenue) * 100
        else:
            growth = 100
            
        # Média de vendas por dia
        days_in_month = (end_of_month - start_of_month).days + 1
        avg_daily_sales = month_revenue / days_in_month if days_in_month > 0 else 0
        
        return render_template(
            'management/dashboard.html',
            last_12_months=last_12_months,
            sales_by_month=sales_by_month,
            top_products=top_products,
            month_revenue=float(month_revenue),
            total_receivables=float(total_receivables),
            received=float(received),
            pending_receivables=float(pending_receivables),
            total_payables=float(total_payables),
            paid=float(paid),
            pending_payables=float(pending_payables),
            estimated_profit=estimated_profit,
            total_overdue=float(total_overdue),
            growth=growth,
            avg_daily_sales=float(avg_daily_sales),
            total_sales=float(total_sales)
        )
    except Exception as e:
        print(f"Erro na rota /gestao: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@management.route('/correcoes')
@login_required
def correcoes_sistema():
    """Renderiza a página de correções do sistema"""
    return render_template('management/correcoes.html')

@management.route('/reports')
@login_required
def reports():
    categories = Category.query.all()
    return render_template('management/reports.html', categories=categories)

@management.route('/sales-report', methods=['POST'])
@login_required
def sales_report():
    try:
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d')
        
        # Ajusta end_date para incluir todo o dia
        end_date = end_date.replace(hour=23, minute=59, second=59)
        
        print(f"\nGerando relatório de vendas")
        print(f"Data inicial: {start_date}")
        print(f"Data final: {end_date}")
        
        # Busca todas as vendas no período com seus itens
        sales = Sale.query.filter(
            Sale.date.between(start_date, end_date)
        ).order_by(Sale.date).all()
        
        print(f"Vendas encontradas no período: {len(sales)}")
        
        # Cria um novo workbook
        wb = Workbook()
        
        # Planilha de Resumo de Vendas
        ws_summary = wb.active
        ws_summary.title = "Resumo de Vendas"
        
        # Estilo para o cabeçalho
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Cabeçalho do Resumo
        headers = ["Data", "Hora", "Cliente", "Total", "Método de Pagamento", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dados do Resumo
        for row, sale in enumerate(sales, 2):
            ws_summary.cell(row=row, column=1, value=sale.date.strftime('%d/%m/%Y'))
            ws_summary.cell(row=row, column=2, value=sale.date.strftime('%H:%M'))
            ws_summary.cell(row=row, column=3, value=sale.customer.name if sale.customer else "Cliente não identificado")
            ws_summary.cell(row=row, column=4, value=float(sale.total))
            ws_summary.cell(row=row, column=5, value=sale.payment_method)
            ws_summary.cell(row=row, column=6, value=sale.status)
            
            # Cria uma planilha para os itens desta venda
            ws_items = wb.create_sheet(f"Venda {sale.id}")
            
            # Informações da venda
            ws_items.cell(row=1, column=1, value="Informações da Venda")
            ws_items.cell(row=1, column=1).font = Font(bold=True)
            
            ws_items.cell(row=2, column=1, value="Data:")
            ws_items.cell(row=2, column=2, value=sale.date.strftime('%d/%m/%Y %H:%M'))
            
            ws_items.cell(row=3, column=1, value="Cliente:")
            ws_items.cell(row=3, column=2, value=sale.customer.name if sale.customer else "Cliente não identificado")
            
            ws_items.cell(row=4, column=1, value="Total:")
            ws_items.cell(row=4, column=2, value=float(sale.total))
            
            ws_items.cell(row=5, column=1, value="Método de Pagamento:")
            ws_items.cell(row=5, column=2, value=sale.payment_method)
            
            ws_items.cell(row=6, column=1, value="Status:")
            ws_items.cell(row=6, column=2, value=sale.status)
            
            # Cabeçalho dos itens
            headers = ["Produto", "Quantidade", "Preço Unitário", "Desconto", "Subtotal"]
            for col, header in enumerate(headers, 1):
                cell = ws_items.cell(row=8, column=col)
                cell.value = header
                cell.font = header_style
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Dados dos itens
            for item_row, item in enumerate(sale.items, 9):
                ws_items.cell(row=item_row, column=1, value=item.product.name)
                ws_items.cell(row=item_row, column=2, value=float(item.quantity))
                ws_items.cell(row=item_row, column=3, value=float(item.price))
                ws_items.cell(row=item_row, column=4, value=float(item.discount))
                ws_items.cell(row=item_row, column=5, value=float(item.subtotal))
        
        # Ajusta largura das colunas para todas as planilhas
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
        
        # Salva o arquivo
        filename = f"relatorio_vendas_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports', filename)
        
        # Cria o diretório se não existir
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Erro ao gerar relatório: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@management.route('/products-report', methods=['POST'])
@login_required
def products_report():
    try:
        category_id = request.form.get('category')
        low_stock = request.form.get('low_stock') == 'on'
        
        query = Product.query
        
        if category_id:
            query = query.filter(Product.category_id == category_id)
        
        if low_stock:
            query = query.filter(Product.stock <= Product.min_stock)
        
        products = query.all()
        
        print(f"\nGerando relatório de produtos")
        print(f"Categoria: {category_id}")
        print(f"Apenas baixo estoque: {low_stock}")
        print(f"Produtos encontrados: {len(products)}")
        
        # Cria um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Produtos"
        
        # Estilo para o cabeçalho
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Cabeçalho
        headers = ["Nome", "Quantidade", "Quantidade Mínima", "Preço", "Categoria"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for row, p in enumerate(products, 2):
            ws.cell(row=row, column=1, value=p.name)
            ws.cell(row=row, column=2, value=float(p.stock))
            ws.cell(row=row, column=3, value=float(p.min_stock))
            ws.cell(row=row, column=4, value=float(p.selling_price))
            ws.cell(row=row, column=5, value=p.category.name if p.category else 'Sem categoria')
        
        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Salva o arquivo
        filename = f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports', filename)
        
        # Cria o diretório se não existir
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Erro ao gerar relatório de produtos: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@management.route('/financial-report', methods=['POST'])
@login_required
def financial_report():
    try:
        print("\n=== INÍCIO DO RELATÓRIO FINANCEIRO ===")
        
        # Cria um novo workbook
        wb = Workbook()
        
        # Cria planilha de recebíveis
        ws_receivables = wb.active
        ws_receivables.title = 'Contas a Receber'
        
        # Cabeçalho dos recebíveis
        headers = ['Cliente', 'Descrição', 'Valor Total', 'Valor Pago', 'Valor Restante', 'Vencimento', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws_receivables.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Busca todos os recebíveis
        print("\nBuscando recebíveis...")
        receivables = Receivable.query.order_by(Receivable.due_date.asc()).all()
        print(f"Total de recebíveis: {len(receivables)}")
        
        # Adiciona recebíveis
        if receivables:
            for i, r in enumerate(receivables, 2):
                print(f"Adicionando recebível {i-1} de {len(receivables)}")
                print(f"ID: {r.id}")
                print(f"Cliente: {r.customer.name if r.customer else 'N/A'}")
                print(f"Vencimento: {r.due_date}")
                print(f"Valor: {r.amount}")
                print("---")
                
                ws_receivables.cell(row=i, column=1).value = r.customer.name if r.customer else 'N/A'
                ws_receivables.cell(row=i, column=2).value = r.description or 'None'
                ws_receivables.cell(row=i, column=3).value = float(r.amount)
                ws_receivables.cell(row=i, column=4).value = float(r.paid_amount)
                ws_receivables.cell(row=i, column=5).value = float(r.remaining_amount)
                ws_receivables.cell(row=i, column=6).value = r.due_date.strftime('%d/%m/%Y')
                ws_receivables.cell(row=i, column=7).value = r.status
            
            # Linha de total
            total_row = len(receivables) + 2
            total_amount = sum(r.amount for r in receivables)
            total_paid = sum(r.paid_amount for r in receivables)
            total_remaining = sum(r.remaining_amount for r in receivables)
            
            ws_receivables.cell(row=total_row, column=1).value = 'TOTAL'
            ws_receivables.cell(row=total_row, column=1).font = Font(bold=True)
            
            ws_receivables.cell(row=total_row, column=3).value = float(total_amount)
            ws_receivables.cell(row=total_row, column=3).font = Font(bold=True)
            
            ws_receivables.cell(row=total_row, column=4).value = float(total_paid)
            ws_receivables.cell(row=total_row, column=4).font = Font(bold=True)
            
            ws_receivables.cell(row=total_row, column=5).value = float(total_remaining)
            ws_receivables.cell(row=total_row, column=5).font = Font(bold=True)
        else:
            ws_receivables.cell(row=2, column=1).value = "Nenhum recebível encontrado"
        
        # Ajusta largura das colunas
        for col in range(1, len(headers) + 1):
            ws_receivables.column_dimensions[get_column_letter(col)].width = 15
        
        # Cria planilha de contas a pagar
        ws_payables = wb.create_sheet(title='Contas a Pagar')
        
        # Cabeçalho das contas a pagar
        headers = ['Fornecedor', 'Descrição', 'Valor', 'Valor Pago', 'Valor Restante', 'Vencimento', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws_payables.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Busca todas as contas a pagar
        print("\nBuscando contas a pagar...")
        payables = Payable.query.order_by(Payable.due_date.asc()).all()
        print(f"Total de contas a pagar: {len(payables)}")
        
        # Adiciona contas a pagar
        if payables:
            for i, p in enumerate(payables, 2):
                print(f"Adicionando conta a pagar {i-1} de {len(payables)}")
                print(f"ID: {p.id}")
                print(f"Fornecedor: {p.supplier.name if p.supplier else 'N/A'}")
                print(f"Vencimento: {p.due_date}")
                print(f"Valor: {p.amount}")
                print("---")
                
                ws_payables.cell(row=i, column=1).value = p.supplier.name if p.supplier else 'N/A'
                ws_payables.cell(row=i, column=2).value = p.description or ''
                ws_payables.cell(row=i, column=3).value = float(p.amount)
                ws_payables.cell(row=i, column=4).value = float(p.paid_amount)
                ws_payables.cell(row=i, column=5).value = float(p.remaining_amount)
                ws_payables.cell(row=i, column=6).value = p.due_date.strftime('%d/%m/%Y')
                ws_payables.cell(row=i, column=7).value = p.status
            
            # Linha de total
            total_row = len(payables) + 2
            total_amount = sum(p.amount for p in payables)
            total_paid = sum(p.paid_amount for p in payables)
            total_remaining = sum(p.remaining_amount for p in payables)
            
            ws_payables.cell(row=total_row, column=1).value = 'TOTAL'
            ws_payables.cell(row=total_row, column=1).font = Font(bold=True)
            
            ws_payables.cell(row=total_row, column=3).value = float(total_amount)
            ws_payables.cell(row=total_row, column=3).font = Font(bold=True)
            
            ws_payables.cell(row=total_row, column=4).value = float(total_paid)
            ws_payables.cell(row=total_row, column=4).font = Font(bold=True)
            
            ws_payables.cell(row=total_row, column=5).value = float(total_remaining)
            ws_payables.cell(row=total_row, column=5).font = Font(bold=True)
        else:
            ws_payables.cell(row=2, column=1).value = "Nenhuma conta a pagar encontrada"
        
        # Ajusta largura das colunas
        for col in range(1, len(headers) + 1):
            ws_payables.column_dimensions[get_column_letter(col)].width = 15
        
        # Salva o arquivo
        filename = f'relatorio_financeiro_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = os.path.join('static', 'reports', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
        
    except Exception as e:
        print(f"Erro: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
